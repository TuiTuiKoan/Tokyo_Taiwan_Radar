"""
jnto_visitors.py — Pull JNTO 訪日外客統計（台灣月別）into external_stats_taiwan_visitors.

Source: https://www.jnto.go.jp/statistics/data/visitors-statistics/
License: 公共データ利用規約 1.0 (jp-gov-pdl-1.0)
Format: XLSX (one sheet per year, Taiwan in row 9, monthly data in odd columns)

Usage:
    python -m external_stats.jnto_visitors --year 2026 [--dry-run]
    python -m external_stats.jnto_visitors --year 2026 --month 3 [--dry-run]
"""
from __future__ import annotations

import argparse
import io
import logging
import os
import re
import sys

import requests
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from dotenv import load_dotenv
from supabase import create_client

from external_stats.base import ExternalStatsBase

load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
logger = logging.getLogger(__name__)

_JNTO_STATS_URL = "https://www.jnto.go.jp/statistics/data/visitors-statistics/"
_TAIWAN_ROW = 9       # 1-indexed row in each year sheet
_MONTH_COLS = [3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25]   # Jan–Dec
_YOY_COLS   = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26]  # matching YoY %


class JntoVisitorsPuller(ExternalStatsBase):
    source_name = "jnto"
    license_code = "jp-gov-pdl-1.0"

    def _table_name(self):
        return "external_stats_taiwan_visitors"

    # ------------------------------------------------------------------
    # Step 1: Discover the latest XLSX URL from the statistics page
    # ------------------------------------------------------------------
    def _discover_xlsx_url(self) -> str:
        resp = requests.get(
            _JNTO_STATS_URL,
            headers={"User-Agent": "Mozilla/5.0 (TokyoTaiwanRadar; research)"},
            timeout=30,
        )
        resp.raise_for_status()
        # Find all .xlsx hrefs
        hrefs = re.findall(r'href="(/statistics/data/_files/[^"]+\.xlsx)"', resp.text)
        if not hrefs:
            raise RuntimeError("Could not find XLSX link on JNTO statistics page")
        # The first href is the latest file
        return "https://www.jnto.go.jp" + hrefs[0]

    # ------------------------------------------------------------------
    # Step 2: Download XLSX bytes
    # ------------------------------------------------------------------
    def fetch(self, **kwargs) -> bytes:
        url = self._discover_xlsx_url()
        logger.info("Downloading JNTO XLSX: %s", url)
        resp = requests.get(
            url,
            headers={"User-Agent": "Mozilla/5.0 (TokyoTaiwanRadar; research)"},
            timeout=60,
        )
        resp.raise_for_status()
        self._source_url = url
        return resp.content

    # ------------------------------------------------------------------
    # Step 3: Parse XLSX into normalised records
    # ------------------------------------------------------------------
    def parse(self, raw: bytes, year: int, month: int | None = None, **kwargs) -> list[dict]:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        sheet_name = str(year)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in JNTO XLSX. Available: {wb.sheetnames}")
        ws = wb[sheet_name]

        records = []
        months_to_pull = [month] if month else list(range(1, 13))

        for mo in months_to_pull:
            col_idx = _MONTH_COLS[mo - 1]   # openpyxl is 1-indexed
            yoy_col = _YOY_COLS[mo - 1]
            total = ws.cell(row=_TAIWAN_ROW, column=col_idx).value
            yoy = ws.cell(row=_TAIWAN_ROW, column=yoy_col).value
            if total is None:
                # Month not yet published
                continue
            year_month = f"{year:04d}-{mo:02d}"
            yoy_val = round(float(yoy), 2) if yoy is not None else None
            # Cap at NUMERIC(6,2) limit — extreme COVID-recovery YoY (e.g. 50000%) stored as NULL
            if yoy_val is not None and abs(yoy_val) >= 9999.99:
                yoy_val = None
            records.append({
                "year_month": year_month,
                "source": self.source_name,
                "total_visitors": int(total),
                "yoy_change_pct": yoy_val,
                "raw_data": {"row": _TAIWAN_ROW, "col": col_idx, "value": total},
                "source_url": getattr(self, "_source_url", _JNTO_STATS_URL),
                "license_code": self.license_code,
            })
            logger.info("  %s: %s visitors (YoY %s%%)", year_month, total,
                        round(float(yoy), 1) if yoy else "N/A")
        return records


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _get_client():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set in scraper/.env")
    return create_client(url, key)


def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    parser = argparse.ArgumentParser(description="Pull JNTO 訪日台灣客統計")
    parser.add_argument("--year", type=int, required=True, help="Target year (e.g. 2026)")
    parser.add_argument("--month", type=int, default=None, help="Optional: single month 1-12")
    parser.add_argument("--dry-run", action="store_true", help="Print records without upserting")
    args = parser.parse_args()

    sb = _get_client()
    puller = JntoVisitorsPuller(sb, dry_run=args.dry_run)
    raw = puller.fetch()
    records = puller.parse(raw, year=args.year, month=args.month)
    n = puller.upsert(puller._table_name(), records)
    print(f"Done: {n} rows upserted to external_stats_taiwan_visitors")


if __name__ == "__main__":
    main()
